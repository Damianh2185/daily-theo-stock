import streamlit as st
import pandas as pd
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from datetime import datetime
import math

# ── Configuración de la página ──────────────────────────────
st.set_page_config(
    page_title="INVENTARIO DIARIO Y COMPRAS",
    page_icon="🥬",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Estilos CSS Personalizados (Aesthetics) ────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&family=Outfit:wght@400;700;800&display=swap');

    :root {
        --primary: #2ECC71;
        --secondary: #3498DB;
        --bg-dark: #0E1117;
        --card-bg: #1A1C24;
        --text-muted: #8E9AAF;
        --border-color: #2D3139;
    }

    /* Fondo Global y Contenedores */
    .stApp, [data-testid="stHeader"], [data-testid="stAppViewContainer"] {
        background-color: var(--bg-dark) !important;
        color: #FFFFFF;
        font-family: 'Inter', sans-serif;
    }
    
    .main { background-color: var(--bg-dark); }
    
    /* Header */
    .app-header {
        text-align: center;
        padding: 0.8rem 1rem;
        background: linear-gradient(135deg, #1A1C24 0%, #12141D 100%);
        border-radius: 12px;
        margin-bottom: 1.5rem;
        border: 1px solid var(--border-color);
        box-shadow: 0 8px 16px rgba(0,0,0,0.4);
    }
    .app-header h1 {
        margin: 0;
        font-family: 'Outfit', sans-serif;
        font-size: 1.3rem;
        font-weight: 700;
        letter-spacing: -1px;
        background: linear-gradient(to right, #2ECC71, #3498DB);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .app-header p { color: var(--text-muted); font-size: 0.85rem; margin-top: 0.2rem; font-weight: 400; }
    
    /* Pestañas (Tabs) */
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 35px;
        background-color: #1A1C24;
        border-radius: 6px 6px 0 0;
        color: var(--text-muted);
        border: 1px solid var(--border-color);
        padding: 0 12px;
        font-size: 0.8rem;
        transition: all 0.3s;
    }
    .stTabs [aria-selected="true"] {
        background-color: var(--primary) !important;
        color: white !important;
        border-color: var(--primary) !important;
    }

    /* Cards */
    .upload-card {
        background: var(--card-bg);
        border: 1px solid var(--border-color);
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        transition: all 0.3s ease;
    }
    .upload-card:hover { 
        border-color: var(--primary); 
        box-shadow: 0 4px 12px rgba(46, 204, 113, 0.1); 
    }
    
    .card-header { display: flex; align-items: center; gap: 0.8rem; margin-bottom: 0.8rem; }
    .card-header h3 { font-size: 1.1rem; margin: 0; }
    .card-icon { font-size: 1.2rem; padding: 0.5rem; border-radius: 8px; }
    .card-icon.blue { background: rgba(52, 152, 219, 0.1); color: #3498DB; }
    .card-icon.green { background: rgba(46, 204, 113, 0.1); color: #2ECC71; }
    
    /* Botones - Visibilidad de Letras */
    div.stButton > button, div.stDownloadButton > button, [data-testid="stFileUploader"] button {
        background-color: var(--primary) !important;
        color: #FFFFFF !important;
        border-radius: 8px !important;
        border: none !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        transition: all 0.2s !important;
    }
    div.stButton > button:hover, div.stDownloadButton > button:hover, [data-testid="stFileUploader"] button:hover {
        background-color: var(--secondary) !important;
        box-shadow: 0 8px 16px rgba(52, 152, 219, 0.3) !important;
        transform: translateY(-1px);
        color: #FFFFFF !important;
    }

    /* Asegurar visibilidad en el uploader */
    [data-testid="stFileUploader"] {
        background-color: rgba(255, 255, 255, 0.05);
        border-radius: 12px;
        padding: 10px;
    }
    [data-testid="stFileUploader"] label {
        color: white !important;
    }

    /* Checkbox visible en tema oscuro */
    .stCheckbox label, .stCheckbox span, .stCheckbox div {
        color: #FFFFFF !important;
    }
    .stCheckbox label {
        font-weight: 600 !important;
    }

    /* Métricas Reutilizables */
    .metric-container { display: flex; gap: 0.8rem; margin: 0.8rem 0; }
    .metric-item {
        flex: 1;
        background: rgba(255,255,255,0.03);
        border-radius: 10px;
        padding: 0.8rem;
        text-align: center;
        border: 1px solid var(--border-color);
    }
    .metric-value { font-size: 1.3rem; font-weight: 800; color: var(--secondary); }
    .metric-label { font-size: 0.8rem; color: var(--text-muted); margin-top: 0.2rem; }
    .metric-value.success { color: var(--primary); }

    /* Footer */
    .app-footer { 
        text-align: center; 
        color: var(--text-muted); 
        padding: 2rem 0; 
        font-size: 0.8rem; 
        border-top: 1px solid var(--border-color); 
        margin-top: 2rem; 
    }
</style>
""", unsafe_allow_html=True)

# ── Datos Predeterminados ──────────────────────────────────
ARTICULOS_DEFAULT = [
    {"Almacen": "CEDIS ARUBAM", "Sub-Categoria": "PANES", "Código": "018507", "Descripción": "BACON PRE-COCIDO REBANADO"},
]

VEGETALES_DEFAULT = [
    {"Clave": "001015", "Producto": "Arandano Fresco"},
    {"Clave": "001026", "Producto": "Fresa"},
    {"Clave": "001052", "Producto": "Melón"},
    {"Clave": "001061", "Producto": "Papaya"},
    {"Clave": "001072", "Producto": "Guineo Amarillo"},
    {"Clave": "001091", "Producto": "Cebollin"},
    {"Clave": "001093", "Producto": "Cilantro"},
    {"Clave": "001110", "Producto": "Perejil Chino"},
    {"Clave": "001115", "Producto": "Ajo Pelado"},
    {"Clave": "001204", "Producto": "Apio"},
    {"Clave": "001221", "Producto": "Cebolla Blanca Jumbo"},
    {"Clave": "001223", "Producto": "Cebolla Morada"},
    {"Clave": "001230", "Producto": "Chayote"},
    {"Clave": "001259", "Producto": "Mazorca Pelada"},
    {"Clave": "001270", "Producto": "Jitomate Cherry"},
    {"Clave": "001282", "Producto": "Lechuga Romana"},
    {"Clave": "001313", "Producto": "Zanahoria"},
    {"Clave": "015932", "Producto": "Aguacate Papelillo"},
    {"Clave": "015971", "Producto": "Culantro"},
    {"Clave": "016041", "Producto": "Limón Tahití"},
    {"Clave": "016074", "Producto": "Papa Criolla"},
    {"Clave": "016083", "Producto": "Pepino Nacional"},
    {"Clave": "016092", "Producto": "Platano Maduro"},
    {"Clave": "016107", "Producto": "Repollo Morado"},
    {"Clave": "016140", "Producto": "Tomate 3x3"},
    {"Clave": "016152", "Producto": "Uvas Rojas sin Semilla"},
    {"Clave": "039029", "Producto": "Ñame Fresco"},
    {"Clave": "040992", "Producto": "Aji Pico E Loro"},
]

# Mapeo global de empaques preferidos (codigo -> texto de paquete) - abreviado para caber en columna
EMPAQUES_PREFERIDOS = {
    "001026": "PAQ. DE 300 GR",
    "001015": "PAQ. DE 125 GR",
}

# ── Funciones Auxiliares (Estabilidad Pro) ──────────────────
def limpiar_valor_numerico(valor) -> float:
    """Limpia strings con $, comas y espacios de forma ultra-confiable."""
    if pd.isna(valor) or valor is None: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    s = str(valor).strip().replace("$", "").replace(" ", "").replace(",", "")
    if not s or s.lower() in ["n/a", "no", "-"]: return 0.0
    try: return float(s)
    except:
        # Reintentar si la coma era el separador decimal
        try: return float(str(valor).replace("$","").replace(".","").replace(",", "."))
        except: return 0.0

def encontrar_columna(df: pd.DataFrame, posibles: list) -> str:
    """Busca columna por nombre exacto o parcial."""
    cols = [str(c).strip().upper() for c in df.columns]
    for p in posibles:
        p_up = p.upper()
        if p_up in cols: return df.columns[cols.index(p_up)]
        for real_col in df.columns:
            if p_up in str(real_col).upper(): return real_col
    return ""

def leer_excel(archivo) -> pd.DataFrame:
    """Lectura robusta de Excel con mapeo de Clave y preservación de ceros."""
    try: df = pd.read_excel(archivo, dtype=str, engine="openpyxl")
    except: df = pd.read_excel(archivo, dtype=str)
    
    # Búsqueda flexible de la columna clave
    c_clave = encontrar_columna(df, ["Clave", "Código", "Codigo", "SKU"])
    if c_clave: 
        df = df.rename(columns={c_clave: "Clave"})
        df["Clave"] = df["Clave"].astype(str).str.strip().str.upper()
    return df

def leer_excel_todas_hojas(archivo) -> pd.DataFrame:
    """Carga todas las hojas, normaliza columnas y preserva ceros."""
    try: sheets = pd.read_excel(archivo, sheet_name=None, dtype=str, engine="openpyxl")
    except: sheets = pd.read_excel(archivo, sheet_name=None, dtype=str)
    dfs = []
    for df in sheets.values():
        c_cl = encontrar_columna(df, ["Clave", "Código", "Codigo", "SKU"])
        if c_cl: df = df.rename(columns={c_cl: "Clave"}).assign(Clave=lambda x: x["Clave"].astype(str).str.strip().str.upper())
        c_pr = encontrar_columna(df, ["Producto", "Descripción", "Description"])
        if c_pr: df = df.rename(columns={c_pr: "Producto"}).assign(Producto=lambda x: x["Producto"].astype(str).str.strip())
        c_te = encontrar_columna(df, ["Inventarios Teóricos", "Teórico", "Teorico", "Stock", "merawey"])
        if c_te: df = df.rename(columns={c_te: "Inventarios Teóricos"})
        # Detectar columna de unidad de medida (varias variantes posibles) y normalizarla a 'Unidad'
        c_um = encontrar_columna(df, ["Unidad", "UM", "U/M", "Unidad de Medida", "UNIDAD DE MEDIDA", "UNID", "U.M."])
        if c_um:
            df = df.rename(columns={c_um: "Unidad"}).assign(Unidad=lambda x: x["Unidad"].astype(str).str.strip())
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

# Lista EXPANDIDA de unidades conocidas (con variaciones)
UNIDADES_CONOCIDAS = [
    # Peso
    "KG", "KILO", "KILOS", "KILOGRAMO", "KILOGRAMOS",
    "G", "GR", "GRAMO", "GRAMOS",
    "LB", "LIBRA", "LIBRAS",
    "OZ", "ONZA", "ONZAS",
    # Cantidad
    "PZA", "PIEZA", "PIEZAS", "PZ",
    "CAJA", "CAJAS",
    "PACK", "PAQUETE", "PAQUETES", "PK",
    "BOLSA", "BOLSAS",
    "UNIDAD", "UNIDADES", "UND", "U",
    # Líquido
    "ML", "MILILITRO", "MILILITROS",
    "L", "LITRO", "LITROS",
    # Otros
    "POTE", "LATA", "LATAS", "LT", "DOCENA"
]

def limpiar_texto_pdf(texto: str) -> str:
    """Limpia y normaliza el texto del PDF."""
    # Normalizar espacios y saltos de línea
    texto = re.sub(r'\s+', ' ', texto)
    # Eliminar caracteres especiales problemáticos
    texto = texto.replace('ó', 'o').replace('é', 'e').replace('á', 'a')
    return texto.strip()

def normalizar_unidad(unidad: str) -> str:
    """Normaliza variantes de unidad a formas consistentes.

    - Si la unidad contiene números (ej. "PIEZA 300 GRAMOS") se devuelve tal cual (limpio)
    - Para términos simples devuelve una forma estándar: KG, GRAMOS, PZA, CAJA, UND, ML, L, LB
    """
    if unidad is None:
        return ""
    s = str(unidad).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    if not s:
        return ""
    # Si contiene un número, devolver la especificación completa (por ejemplo PIEZA 300 GRAMOS)
    if re.search(r"\d", s):
        return s
    # Mapear variantes comunes
    if re.search(r"\b(KG|KILO|KILOS|KILOGRAMO|KILOGRAMOS)\b", s):
        return "KG"
    if re.search(r"\b(GR|G|GRAMO|GRAMOS)\b", s):
        return "GRAMOS"
    if re.search(r"\b(PZA|PIEZA|PIEZAS|PZ)\b", s):
        return "PZA"
    if re.search(r"\b(CAJA|CAJAS)\b", s):
        return "CAJA"
    if re.search(r"\b(UND|UNIDAD|UNIDADES|U)\b", s):
        return "UND"
    if re.search(r"\b(ML|MILILITRO|MILILITROS)\b", s):
        return "ML"
    if re.search(r"\b(L|LITRO|LITROS)\b", s):
        return "L"
    if re.search(r"\b(LB|LIBRA|LIBRAS)\b", s):
        return "LB"
    # Si no coincide, devolver la versión limpiada
    return s


def parse_pdf_quantity_to_kg(cantidad_raw, unidad_raw, codigo=None, empaques_preferidos=None):
    """Convertir cantidad+unidad del PDF a kilogramos y detectar tamaño de paquete si aplica.

    Retorna: (sol_kg: float, display_unit: str, package_size_kg: float|None)
    """
    try:
        cantidad = float(cantidad_raw)
    except:
        cantidad = 0.0
    s = (unidad_raw or "").upper().strip()
    package_size_kg = None
    display_unit = s if s else ''

    # Si no hay unidad en PDF, usar mapeo preferido si existe
    if (not s or s == 'N/D') and empaques_preferidos and codigo in empaques_preferidos:
        display_unit = empaques_preferidos[codigo]
        m = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", display_unit, re.IGNORECASE)
        if m:
            size = float(m.group(1).replace(',', '.'))
            u = m.group(2).upper()
            if 'G' in u or 'GRAM' in u:
                package_size_kg = size / 1000.0
            else:
                package_size_kg = size
        # cantidad probablemente representa número de paquetes
        if package_size_kg:
            return cantidad * package_size_kg, display_unit, package_size_kg
        return cantidad, display_unit, None

    # Detectar paquetes/peceras/piezas con tamaño incluido
    if re.search(r"\b(PAQUETE|PAQ|PIEZA|PZA|PACK|CAJA|PK)\b", s):
        m = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", s, re.IGNORECASE)
        if m:
            size = float(m.group(1).replace(',', '.'))
            u = m.group(2).upper()
            if 'G' in u or 'GRAM' in u:
                package_size_kg = size / 1000.0
            else:
                package_size_kg = size
            # cantidad es número de paquetes
            return cantidad * package_size_kg, display_unit or s, package_size_kg
        # paquete sin tamaño explícito
        return cantidad, display_unit or s, None

    # Si la unidad es gramos
    if re.search(r"\b(GR|G|GRAMO|GRAMOS)\b", s):
        return cantidad / 1000.0, display_unit or s, None

    # Si la unidad es kilogramos
    if re.search(r"\b(KG|KILO|KILOGRAMO|KILOGRAMOS)\b", s):
        return cantidad, display_unit or s, None

    # Si la unidad contiene un número y un sufijo
    m = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", s, re.IGNORECASE)
    if m:
        size = float(m.group(1).replace(',', '.'))
        u = m.group(2).upper()
        if 'G' in u or 'GRAM' in u:
            return size / 1000.0, display_unit or s, None
        return size, display_unit or s, None

    # Heurístico: si cantidad es grande (>50), es probable que sea gramos
    if cantidad > 50:
        return cantidad / 1000.0, display_unit or s, None

    return cantidad, display_unit or s, None


def compute_packages_needed(total_kg, package_size_kg):
    if not package_size_kg or package_size_kg <= 0:
        return ''
    return math.ceil(total_kg / package_size_kg)

def extraer_datos_pdf(archivo) -> dict:
    """Extracción MEJORADA de PDF con detección robusta de unidades."""
    datos = {}
    try:
        reader = PdfReader(archivo)
        # Procesar cada página por separado para mejor control
        todas_las_claves = {}
        
        for page_num, page in enumerate(reader.pages):
            texto_pagina = page.extract_text() or ""
            
            # Limpiar texto
            texto_pagina = limpiar_texto_pdf(texto_pagina)
            
            # Buscar números de 5+ dígitos (claves)
            pos_claves = list(re.finditer(r"(\d{5,})", texto_pagina))
            
            for i, m_clave in enumerate(pos_claves):
                clave = m_clave.group(1)
                inicio = m_clave.end()
                fin = pos_claves[i+1].start() if i+1 < len(pos_claves) else len(texto_pagina)
                
                # Segmento de texto asociado a la clave
                segmento = texto_pagina[inicio:fin].strip()
                
                # --- BÚSQUEDA DE CANTIDAD Y UNIDAD ---
                cantidad = 0.0
                unidad = "N/D"

                # Priorizar detección de piezas/paquetes cuando el PDF contiene ambos
                m_piece = re.search(r"(\d+[\.,]?\d*)\s*(PIEZA|PZA|PAQ|PAQUETE|PACK)\b", segmento, re.IGNORECASE)
                if m_piece:
                    try:
                        cantidad = float(m_piece.group(1).replace(',', '.'))
                    except:
                        cantidad = 0.0
                    pos_inicio_unit = m_piece.start(0)
                    resto_desde_unit = segmento[pos_inicio_unit:]
                    match_desc = re.match(r"([A-Za-z0-9\s]{1,150}?)(?:\s+\d+[\.,]\d+|$)", resto_desde_unit)
                    if match_desc:
                        unidad = " ".join(match_desc.group(1).strip().split())[:80]
                    else:
                        unidad = m_piece.group(2).upper()
                else:
                    # Estrategia 1: Buscar cantidad + unidad juntas
                    # Patrón: número (con decimales) + espacio(s) + unidad
                    unidades_sorted = sorted(UNIDADES_CONOCIDAS, key=len, reverse=True)
                    for unit in unidades_sorted:
                        # Búsqueda case-insensitive
                        patron = rf"(\d+[\.,]?\d*)\s*{unit}\b"
                        match = re.search(patron, segmento, re.IGNORECASE)
                        if match:
                            # Extraer cantidad
                            try:
                                cant_str = match.group(1).replace(',', '.')
                                cantidad = float(cant_str)
                            except:
                                cantidad = 0.0
                            # Extraer unidad + descripción (hasta 100 chars o siguiente número)
                            pos_inicio_unit = match.start(0)
                            resto_desde_unit = segmento[pos_inicio_unit:]
                            match_desc = re.match(rf"([A-Za-z0-9\s]{{1,150}}?)(?:\s+\d+[\.,]\d+|$)", resto_desde_unit)
                            if match_desc:
                                desc_raw = match_desc.group(1).strip()
                                unidad = " ".join(desc_raw.split())[:80]
                            else:
                                unidad = unit.upper()
                            break
                
                # Estrategia 2: Si no encontró, buscar solo el número más cercano al inicio
                if unidad == "N/D":
                    match_num = re.search(r"(\d+[\.,]?\d*)", segmento)
                    if match_num:
                        try:
                            cantidad = float(match_num.group(1).replace(",", "."))
                        except:
                            cantidad = 0.0
                
                # Guardar (si esta clave ya existe, mantener la que tenga mejor info)
                if clave not in todas_las_claves or unidad != "N/D":
                    todas_las_claves[clave] = {
                        "cantidad": cantidad,
                        "unidad": unidad.upper()
                    }
        
        datos = todas_las_claves
        st.session_state["last_pdf_text"] = f"Se procesaron {len(datos)} artículos del PDF"
        
    except Exception as e:
        st.error(f"Error al procesar PDF: {e}")
    
    return datos

# ── Generación de Archivos (Estilos) ───────────────────────
def aplicar_estilos_base(ws, df):
    # Estilos profesionales mejorados
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")  # Gris oscuro profesional
    header_font = Font(color="FFFFFF", bold=True, size=14, name="Segoe UI")  # Blanco sobre oscuro
    border_style = Border(
        left=Side(style='thin', color="34495E"),
        right=Side(style='thin', color="34495E"),
        top=Side(style='thin', color="34495E"),
        bottom=Side(style='thin', color="34495E")
    )

    # Aplicar encabezado
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_style

    # Alternar colores de fila y aplicar bordes
    fill_light = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")  # Gris muy claro
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco
    data_font = Font(color="2C3E50", size=11, name="Segoe UI")

    # Mapa de alineación por columna (lowercase keys)
    align_map = {
        'codigo': 'center',
        'código': 'center',
        'clave': 'center',
        'descripcion': 'left',
        'descripción': 'left',
        'inv. merawey': 'right',
        'inventario merawey': 'right',
        'inv merawey': 'right',
        'inventarios teoricos': 'right',
        'solicitado': 'right',
        'solicitado a compras': 'right',
        'um compras': 'center',
        'unidad': 'center',
        'unidad de medida solicitada a compras': 'center',
        'total': 'right',
        'cant. a solicitar': 'right',
        'cantidad a solicitar': 'right',
        'conteo fisico': 'right',
    }

    for row_idx in range(2, len(df) + 2):
        fill = fill_light if (row_idx % 2 == 0) else fill_white
        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = border_style
            c.font = data_font
            # Determinar alineación basada en el nombre de la columna
            col_name = str(df.columns[col_idx - 1]).lower()
            alignment = align_map.get(col_name, None)
            if alignment == 'right':
                c.alignment = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=True)
            elif alignment == 'center':
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

def calcular_ancho_columna(nombre_columna: str, datos_columna=None) -> float:
    """
    Calcula inteligentemente el ancho de la columna basado en:
    - Longitud del encabezado
    - Tipo de datos (números vs texto)
    - Relleno profesional
    
    COLUMNAS NUMÉRICAS: Tienen ancho UNIFORME (16px exacto)
    """
    nombre = str(nombre_columna).strip()
    
    # Anchos definidos manualmente para columnas especiales
    anchos_especiales = {
        "Código": 14,                    # Códigos son siempre cortos (6 dígitos)
        "Descripción": 32,               # Nombres de productos pueden variar
        "Cant. a Solicitar": 32,         # AMPLIADO: igual que descripción para escribir valores
        "Inv. Merawey": 16,              # UNIFORME: columnas numéricas
        "Solicitado": 16,                # UNIFORME: columnas numéricas
        "Total": 16,                     # UNIFORME: columnas numéricas
    }
    
    # Si está en especiales, usarlo directamente
    if nombre in anchos_especiales:
        return anchos_especiales[nombre]
    
    # Para el resto (como UM Compras), calcular dinámicamente
    CHAR_WIDTH = 1.4
    PADDING = 2
    ancho = len(nombre) * CHAR_WIDTH + PADDING
    return max(10, min(50, ancho))

def aplicar_ancho_inteligente(ws, df_export):
    """Aplica ancho inteligente y uniforme a las columnas."""
    for col_idx, col_name in enumerate(df_export.columns, 1):
        ancho = calcular_ancho_columna(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

def generar_excel_vegetales(df, portrait=False):
    # Renombrar columnas a versiones más cortas y profesionales
    df_export = df.copy()
    columnas_renombre = {
        "Código": "Código",
        "Descripción": "Descripción",
        "inventario merawey": "Inv. Merawey",
        "Solicitado a compras": "Solicitado",
        "UNIDAD DE MEDIDA SOLICITADA A COMPRAS": "UM Compras",
        "Total": "Total",
        # Mantener el encabezado completo para que el usuario no lo vea abreviado
        "Cantidad a Solicitar": "Cantidad a Solicitar"
    }
    df_export = df_export.rename(columns=columnas_renombre)

    # Usar mapeo global de empaques preferidos (definido en EMPAQUES_PREFERIDOS)
    empaques_preferidos = EMPAQUES_PREFERIDOS

    # Asegurar columnas necesarias
    if 'UM Compras' not in df_export.columns:
        df_export['UM Compras'] = ''

    # Normalizar texto y aplicar overrides de empaque cuando aplique
    df_export['UM Compras'] = df_export['UM Compras'].fillna('').astype(str).str.strip()
    if 'Código' in df_export.columns:
        for codigo, paquete in empaques_preferidos.items():
            mask = (df_export['Código'].astype(str).str.strip() == codigo)
            if mask.any():
                # Forzar que ciertos productos se soliciten en empaques preferidos (p. ej. fresas y arandanos)
                df_export.loc[mask, 'UM Compras'] = paquete
                # No escribir notas por fila; la instrucción completa estará al final del reporte

    # (No generamos una hoja separada de notas aquí; las instrucciones se colocan
    #  en el área final del mismo 'Reporte' para evitar duplicados.)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Reporte")
        ws = writer.sheets["Reporte"]

        # Aplicar estilos base mejorados
        aplicar_estilos_base(ws, df_export)
        
        # Colores específicos para columnas de datos (MEJORADO)
        cols_colores = {
            "Inv. Merawey": "D6EAF8",      # Azul claro profesional
            "Solicitado": "D5F4E6",         # Verde claro profesional
            "UM Compras": "FADBD8",         # Naranja/salmón claro
            "Total": "F9E79F"               # Amarillo profesional
        }
        
        for col_idx, col_name in enumerate(df_export.columns, 1):
            if col_name in cols_colores:
                fill = PatternFill(start_color=cols_colores[col_name], end_color=cols_colores[col_name], fill_type="solid")
                for row_idx in range(2, len(df_export) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    # Mantener bordes y fuente
                    cell.border = Border(
                        left=Side(style='thin', color="34495E"),
                        right=Side(style='thin', color="34495E"),
                        top=Side(style='thin', color="34495E"),
                        bottom=Side(style='thin', color="34495E")
                    )
                    cell.font = Font(color="2C3E50", size=10, name="Segoe UI")
                    # No cambiar alineación aquí: la función aplicar_estilos_base
                    # ya aplica la alineación profesional por columna.
        
        # Configuración de impresión (REQUERIDO: CARTA/HORIZONTAL)
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT if portrait else ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        # Márgenes de impresión en pulgadas (izq, der, arriba, abajo)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # Altura de filas mejorada
        ws.row_dimensions[1].height = 44 if portrait else 36  # Cabecera con más espacio para evitar cortes
        for row_idx in range(2, len(df_export) + 2):
            ws.row_dimensions[row_idx].height = 26 if portrait else 22  # Filas de datos más espaciosas en vertical
        
        # Aplicar ancho inteligente y uniforme
        aplicar_ancho_inteligente(ws, df_export)

        # Ajustes manuales de ancho para aprovechar mejor el espacio disponible
        preferred_widths = {
            # Código debe tener el mismo ancho que las columnas numéricas C..F
            'Código': 16 if portrait else 20,
            'Descripción': 32 if portrait else 44,
            'Inv. Merawey': 16,
            'Solicitado': 16,
            'UM Compras': 16,
            'Total': 16,
            # Mostrar encabezado completo y dar más espacio a la columna G
            'Cantidad a Solicitar': 30 if portrait else 45
        }
        for col_name, w in preferred_widths.items():
            try:
                idx = list(df_export.columns).index(col_name) + 1
                ws.column_dimensions[get_column_letter(idx)].width = w
            except ValueError:
                continue
        
        # Formato numérico: Inv. Merawey y Total siempre 2 decimales.
        # Para 'Solicitado' mostrar sin decimales si la unidad de compra es por paquete (PAQ),
        # en caso contrario mostrar 2 decimales.
        idx_inv = None
        idx_solicitado = None
        idx_um = None
        idx_total = None
        for i, col_name in enumerate(df_export.columns, 1):
            low = str(col_name).strip().lower()
            if low in ["inv. merawey", "inventario merawey", "inv merawey"]:
                idx_inv = i
            if low in ["solicitado", "solicitado a compras"]:
                idx_solicitado = i
            if low in ["um compras", "unidad de medida solicitada a compras", "unidad"]:
                idx_um = i
            if low == "total":
                idx_total = i

        for row_idx in range(2, len(df_export) + 2):
            if idx_inv:
                ws.cell(row=row_idx, column=idx_inv).number_format = '0.00'
            if idx_total:
                ws.cell(row=row_idx, column=idx_total).number_format = '0.00'
            if idx_solicitado:
                um_val = ws.cell(row=row_idx, column=idx_um).value if idx_um else ''
                um_str = str(um_val or '').upper()
                if 'PAQ' in um_str or 'PAQUETE' in um_str:
                    ws.cell(row=row_idx, column=idx_solicitado).number_format = '0'
                else:
                    ws.cell(row=row_idx, column=idx_solicitado).number_format = '0.00'

        # Agregar nota profesional en la hoja 'Reporte' (área inferior) con instrucciones por producto
        try:
            num_cols = len(df_export.columns)
            start_row = len(df_export) + 4
            note_lines = []
            for codigo, paquete in empaques_preferidos.items():
                mask = df_export['Código'].astype(str).str.strip() == codigo
                if mask.any():
                    descripcion = str(df_export.loc[mask, 'Descripción'].iloc[0])
                    # obtener total en kg
                    try:
                        total_kg = float(df_export.loc[mask, 'Total'].iloc[0])
                    except:
                        total_kg = 0.0
                    # obtener unidad de compra preferida (de la fila si existe)
                    unidad_fila = str(df_export.loc[mask, 'UM Compras'].iloc[0]) if 'UM Compras' in df_export.columns else paquete
                    # determinar tamaño de paquete
                    m = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", unidad_fila, re.IGNORECASE)
                    paquete_text = unidad_fila
                    paquetes_necesarios = ''
                    if m:
                        size = float(m.group(1).replace(',', '.'))
                        u = m.group(2).upper()
                        if 'G' in u or 'GRAM' in u:
                            package_size_kg = size / 1000.0
                        else:
                            package_size_kg = size
                        paquetes_necesarios = compute_packages_needed(total_kg, package_size_kg)
                        paquete_text = f"{int(size)} {m.group(2).upper()}"
                # Nota: solo indicar cómo solicitar en Merawey, sin el total calculado
                note_lines.append(f"{codigo} - {descripcion}: Solicitar en Merawey como '{unidad_fila}'.")

            if note_lines:
                note_text = "NOTAS PARA SOLICITUD EN MERAWEY:\n" + "\n".join(note_lines)
                # Usar sólo 3 filas al final para la nota; combinarlas y centrar
                start_row = len(df_export) + 3
                end_row = start_row + 2  # 3 filas en total
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=num_cols)
                cell = ws.cell(row=start_row, column=1)
                cell.value = note_text
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.font = Font(size=10, color='000000')
                # Aplicar fondo sólo en las 3 filas de la nota
                fill_note = PatternFill(start_color="FFF9E8", end_color="FFF9E8", fill_type="solid")
                for r in range(start_row, end_row + 1):
                    ws.row_dimensions[r].height = 32
                    for c in range(1, num_cols + 1):
                        ws.cell(row=r, column=c).fill = fill_note
        except Exception:
            pass
    
    return output.getvalue()

def generar_excel_inventario(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inventario")
        ws = writer.sheets["Inventario"]
        aplicar_estilos_base(ws, df)
        # Aplicar ancho inteligente
        aplicar_ancho_inteligente(ws, df)
    return output.getvalue()

def generar_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    return output.getvalue()

def obtener_nombre_archivo_vegetales():
    # 24 abr 2026 Teórico de Vegetales
    meses = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
    ahora = datetime.now()
    return f"{ahora.day} {meses[ahora.month-1]} {ahora.year} Teórico de Vegetales.xlsx"

# ── Interfaz Principal ─────────────────────────────────────
def main():
    st.markdown('<div class="app-header"><h1>INVENTARIO DIARIO Y COMPRAS 🥬</h1></div>', unsafe_allow_html=True)
    tabs = st.tabs(["🚀 Automático (Vegetales)", "📊 Inventario", "📝 Manual"])
    
    # --- PESTAÑA VEGETALES ---
    with tabs[0]:
        st.markdown('<div class="upload-card"><div class="card-header"><div class="card-icon green">🥬</div><h3>Reporte Teórico de Vegetales</h3></div><p>Cruce de Inventario Merawey + Orden de Compra PDF</p></div>', unsafe_allow_html=True)
        v_inv = st.file_uploader("1. Inventario Principal (Excel)", type=["xlsx", "xls"], key="v1")
        v_pdf = st.file_uploader("2. Orden de Compra (PDF Opcional)", type=["pdf"], key="v2")
        export_portrait = st.checkbox(
            "Exportar archivo de excel en vertical.",
            value=False,
            help="Selecciona para generar el archivo en orientación vertical manteniendo los mismos datos y estilos."
        )
        
        if st.button("🚀 Procesar Reporte Vegetales") and v_inv:
            with st.spinner("Procesando..."):
                try:
                    df_full = leer_excel_todas_hojas(v_inv)
                    pdf_data = extraer_datos_pdf(v_pdf) if v_pdf else {}
                    
                    filas = []
                    enc = 0
                    for v in VEGETALES_DEFAULT:
                        match = df_full[(df_full["Clave"] == v["Clave"]) | (df_full["Producto"].str.contains(v["Producto"], na=False, case=False))]
                        teo = limpiar_valor_numerico(match.iloc[0]["Inventarios Teóricos"]) if not match.empty else 0.0
                        if not match.empty: enc += 1
                        
                        d_pdf = pdf_data.get(v["Clave"], {})
                        # Parsear cantidad/unidad del PDF y convertir a KG
                        uni_pdf_raw = str(d_pdf.get("unidad", "") or "").strip()
                        cant_pdf_raw = d_pdf.get("cantidad", 0.0)
                        sol_kg, display_unit, package_size_kg = parse_pdf_quantity_to_kg(cant_pdf_raw, uni_pdf_raw, v["Clave"], EMPAQUES_PREFERIDOS)

                        # Si no se detectó unidad desde PDF, intentar desde el inventario (match)
                        if (not display_unit or display_unit.upper() in ['', 'N/D']) and not match.empty:
                            if 'Unidad' in match.columns and pd.notna(match.iloc[0]['Unidad']) and str(match.iloc[0]['Unidad']).strip() != '':
                                sol_kg_from_inv, display_unit_from_inv, pkg_from_inv = parse_pdf_quantity_to_kg(0, match.iloc[0]['Unidad'], v["Clave"], EMPAQUES_PREFERIDOS)
                                # keep package info if available
                                if display_unit_from_inv and display_unit_from_inv.upper() not in ['', 'N/D']:
                                    display_unit = display_unit_from_inv
                                    if pkg_from_inv:
                                        package_size_kg = pkg_from_inv

                        # Si hay mapeo de empaque preferido, usarlo para mostrar (abreviado)
                        if v["Clave"] in EMPAQUES_PREFERIDOS:
                            display_unit = EMPAQUES_PREFERIDOS[v["Clave"]]

                        # Normalizar display unit para mostrar en el Excel
                        uni = normalizar_unidad(display_unit) if display_unit else 'N/D'

                        # Determinar cantidad reportada en el PDF (si existe)
                        try:
                            pdf_qty = float(cant_pdf_raw) if cant_pdf_raw is not None else None
                        except:
                            pdf_qty = None

                        # Si no tenemos tamaño de paquete detectado, intentar extraerlo del display_unit
                        if package_size_kg is None:
                            # Buscar dentro de display_unit
                            m_size = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", str(display_unit or ''), re.IGNORECASE)
                            # Si no se encontró en display_unit, buscar en el mapeo global EMPAQUES_PREFERIDOS
                            if not m_size and v["Clave"] in EMPAQUES_PREFERIDOS:
                                m_size = re.search(r"(\d+[\.,]?\d*)\s*(G|GRAMOS|GRAMO|GR|KG|KILO)", EMPAQUES_PREFERIDOS[v["Clave"]], re.IGNORECASE)
                            if m_size:
                                size = float(m_size.group(1).replace(',', '.'))
                                u = m_size.group(2).upper()
                                if 'G' in u or 'GRAM' in u:
                                    package_size_kg = size / 1000.0
                                else:
                                    package_size_kg = size

                        # Calcular sol_kg final en kilogramos
                        sol_kg_final = sol_kg
                        if package_size_kg is not None:
                            if pdf_qty is not None:
                                sol_kg_final = float(pdf_qty) * package_size_kg
                            else:
                                # si pdf_qty no está, intentar derivar desde sol_kg recibido
                                try:
                                    sol_kg_final = float(sol_kg) if sol_kg else 0.0
                                except:
                                    sol_kg_final = float(sol_kg)

                        # Calcular total definitivo en KG (teórico + solicitado en kg)
                        total_kg = teo + sol_kg_final

                        # Preparar valor visible en columna 'Solicitado a compras'
                        if package_size_kg is not None:
                            # Mostrar número de paquetes (preferir el valor tal cual en el PDF si existe)
                            if pdf_qty is not None:
                                solicitado_count = pdf_qty
                            else:
                                solicitado_count = sol_kg_final / package_size_kg if package_size_kg else 0
                            if abs(round(solicitado_count) - solicitado_count) < 1e-9:
                                solicitado_display = int(round(solicitado_count))
                            else:
                                solicitado_display = round(solicitado_count, 2)
                        else:
                            solicitado_display = round(sol_kg_final, 2)

                        # NO llenar la columna 'Cantidad a Solicitar' — el usuario la completará manualmente
                        filas.append({
                            "Código": v["Clave"], "Descripción": v["Producto"], "inventario merawey": teo,
                            "Solicitado a compras": solicitado_display, "UNIDAD DE MEDIDA SOLICITADA A COMPRAS": display_unit or 'N/D',
                            "Total": round(total_kg, 2), "Cantidad a Solicitar": ""
                        })
                    res = pd.DataFrame(filas)
                    st.success(f"✅ Reporte generado. Coincidencias en inventario: {enc}")
                    st.dataframe(res, use_container_width=True)
                    
                    nombre = obtener_nombre_archivo_vegetales()
                    st.download_button(
                        "📥 Descargar Reporte Vegetales",
                        generar_excel_vegetales(res, portrait=export_portrait),
                        file_name=nombre
                    )
                except Exception as e:
                    st.error(f"Error: {e}")

    # --- PESTAÑA INVENTARIO ---
    with tabs[1]:
        st.markdown('<div class="upload-card"><div class="card-header"><div class="card-icon blue">📋</div><h3>Formato de Inventario CEDIS</h3></div><p>Genera la hoja de toma física para el almacén</p></div>', unsafe_allow_html=True)
        i_inv = st.file_uploader("Archivo principal de productos", type=["xlsx"], key="i1")
        if st.button("🚀 Generar Formato") and i_inv:
            try:
                df_p = leer_excel(i_inv)
                # Mapear teóricos
                c_t = encontrar_columna(df_p, ["Inventarios Teóricos", "Teórico", "Stock"])
                inv_map = df_p.set_index("Clave")[c_t].to_dict() if c_t else {}
                
                filas = []
                for a in ARTICULOS_DEFAULT:
                    cod = str(a["Código"]).strip()
                    filas.append({
                        "Almacen": a["Almacen"], "Sub-Categoria": a["Sub-Categoria"], "Código": cod,
                        "Descripción": a["Descripción"], "conteo fisico": "", "inventario merawey": inv_map.get(cod, ""),
                        "diferencia": "", "Observaciones": "", "responsable del conteo": ""
                    })
                res = pd.DataFrame(filas)
                st.dataframe(res, use_container_width=True)
                st.download_button("📥 Descargar Formato", generar_excel_inventario(res), file_name="inventario_cedis.xlsx")
            except Exception as e: st.error(f"Error: {e}")

    # --- PESTAÑA MANUAL ---
    with tabs[2]:
        st.markdown('<div class="upload-card"><div class="card-header"><div class="card-icon blue">🔍</div><h3>Cruce Manual</h3></div><p>Cruce libre entre dos archivos usando la columna Clave</p></div>', unsafe_allow_html=True)
        m_p = st.file_uploader("Archivo de Datos (Principal)", type=["xlsx"], key="m1")
        m_c = st.file_uploader("Archivo de Claves (Filtro)", type=["xlsx"], key="m2")
        if st.button("🚀 Ejecutar Cruce") and m_p and m_c:
            try:
                with st.spinner("Procesando cruce..."):
                    df1, df2 = leer_excel(m_p), leer_excel(m_c)
                    if "Clave" in df1.columns and "Clave" in df2.columns:
                        lista_claves = df2["Clave"].dropna().unique().tolist()
                        res = df1[df1["Clave"].isin(lista_claves)].copy()
                        
                        # --- MÉTRICAS INTEGRADAS ---
                        st.markdown(f"""
                        <div class="metric-container">
                            <div class="metric-item">
                                <div class="metric-value">{len(df1):,}</div>
                                <div class="metric-label">Productos Originales</div>
                            </div>
                            <div class="metric-item">
                                <div class="metric-value success">{len(lista_claves):,}</div>
                                <div class="metric-label">Claves a Buscar</div>
                            </div>
                            <div class="metric-item">
                                <div class="metric-value success">{len(res):,}</div>
                                <div class="metric-label">Encontrados</div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if not res.empty:
                            st.success(f"✅ Se encontraron **{len(res)}** productos.")
                            st.dataframe(res, use_container_width=True)
                            st.download_button("📥 Descargar Resultados", generar_excel(res), file_name="cruce_manual.xlsx")
                            
                            # --- CLAVES NO ENCONTRADAS ---
                            claves_encontradas = set(res["Clave"].unique())
                            claves_no_encontradas = [c for c in lista_claves if c not in claves_encontradas]
                            if claves_no_encontradas:
                                with st.expander(f"⚠️ {len(claves_no_encontradas)} claves no encontradas"):
                                    st.dataframe(pd.DataFrame({"Clave no encontrada": claves_no_encontradas}), use_container_width=True)
                        else:
                            st.warning("⚠️ No se encontraron coincidencias.")
                    else:
                        st.error("❌ Falta la columna 'Clave' en uno de los archivos. El sistema la buscó como 'Clave', 'Código' o 'SKU'.")
            except Exception as e: st.error(f"Error: {e}")

    st.markdown('<div class="app-footer">Inventario Pro · Hecho con ❤️ para una gestión impecable</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
